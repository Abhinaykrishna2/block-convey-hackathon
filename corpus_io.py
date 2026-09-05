"""Read supplied evidence without executing macros, formulas or document instructions.

No application code. Source text is separate from curator interpretations. Office
locators refer to XML paragraphs or spreadsheet cells, never guessed page numbers.
"""
from __future__ import annotations

import hashlib
import json
import posixpath
import re
import warnings
from datetime import date, datetime
from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

ROOT = Path(__file__).resolve().parent
NS = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
      'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
SOURCE_EXTENSIONS = {'.docx', '.xlsx', '.pdf', '.png'}


def raw_paths():
    return sorted(p for p in (ROOT / 'Hackathon').rglob('*') if p.is_file() and p.suffix.lower() in SOURCE_EXTENSIONS)


def sha(path):
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def relative(path):
    return Path(path).relative_to(ROOT).as_posix()


def scalar(value):
    return value.isoformat() if isinstance(value, (datetime, date)) else value


def block(source, locator, text, **metadata):
    return {'id': hashlib.sha256((source + '|' + locator).encode()).hexdigest()[:24],
            'source': source, 'locator': locator, 'text': text, **metadata}


def role(path):
    name = Path(path).name
    if name.endswith('.xlsx'):
        return 'questionnaire_template' if 'Questionnaire' in name else 'operational_record'
    if name in ('Regodit_Incident_Management_Policy_v1.0.docx', 'Regodit_code_of_conduct_policy_v1.0.docx', 'Regodit_hr_policy_v1.0.docx', 'Regodit._risk_management_policy_v1.0.docx', 'Regodit_Vendor_Risk_Management_Policy.docx'):
        return 'policy_with_incomplete_approval_metadata'
    if name in ('Employment Contract 01.docx', 'Master Services Agreement.docx', 'Secure Development Lifecycle Document 01.docx', 'BCP_DR_Plan_Solsphere.docx'):
        return 'template_with_placeholders'
    if name == 'Regodit AI_SOC2_Type_II_Report_Test.docx':
        return 'mixed_document_requires_confirmation'
    if '/2. Company policies/' in str(path):
        return 'documented_policy'
    if name == 'VAPT Report 01.docx':
        return 'dated_assessment_report'
    if name == 'Solsphere W-9.pdf':
        return 'unsigned_form_and_generic_instructions'
    return 'diagram_or_reference_requires_confirmation'


def word_blocks(path):
    source = relative(path)
    with ZipFile(path) as archive:
        parts = ['word/document.xml'] + sorted(n for n in archive.namelist()
            if re.fullmatch(r'word/(?:header\d+|footer\d+|footnotes|endnotes|comments)\.xml', n))
        for part in parts:
            tree = ET.fromstring(archive.read(part))
            parents = {c: p for p in tree.iter() for c in p}
            relationships = {}
            rel_path = posixpath.join(posixpath.dirname(part), '_rels', posixpath.basename(part) + '.rels')
            if rel_path in archive.namelist():
                relationships = {e.get('Id'): e.get('Target') for e in ET.fromstring(archive.read(rel_path))}
            paragraph_numbers = {id(p): i for i, p in enumerate(tree.findall('.//w:p', NS), 1)}
            table_numbers = {id(t): i for i, t in enumerate(tree.findall('.//w:tbl', NS), 1)}
            for p in tree.findall('.//w:p', NS):
                pieces = []
                for e in p.iter():
                    if e.tag in (f"{{{NS['w']}}}t", f"{{{NS['w']}}}delText"):
                        pieces.append(e.text or '')
                    elif e.tag == f"{{{NS['w']}}}tab":
                        pieces.append('\t')
                    elif e.tag in (f"{{{NS['w']}}}br", f"{{{NS['w']}}}cr"):
                        pieces.append('\n')
                text = ''.join(pieces)
                fields = [e.text or '' for e in p.findall('.//w:instrText', NS)]
                links = [relationships[e.get(f"{{{NS['r']}}}id")] for e in p.findall('.//w:hyperlink', NS)
                         if e.get(f"{{{NS['r']}}}id") in relationships]
                if not text and not fields and not links:
                    continue
                metadata = {'kind': 'word_paragraph', 'document_role': role(path)}
                if fields:
                    metadata['field_instructions'] = fields
                if links:
                    metadata['hyperlink_targets'] = links
                if p.find('w:pPr/w:numPr', NS) is not None:
                    metadata['list_item'] = True
                ancestor = p
                while ancestor in parents:
                    ancestor = parents[ancestor]
                    if ancestor.tag == f"{{{NS['w']}}}tc":
                        row = parents[ancestor]
                        table = parents[row]
                        metadata['table'] = table_numbers[id(table)]
                        metadata['row'] = table.findall('w:tr', NS).index(row) + 1
                        metadata['cell'] = row.findall('w:tc', NS).index(ancestor) + 1
                        break
                yield block(source, f'{part}#p{paragraph_numbers[id(p)]:05d}', text, **metadata)


def sheet_blocks(path):
    from openpyxl import load_workbook
    from openpyxl.worksheet.formula import ArrayFormula
    source = relative(path)
    # This is read-only. Preserve unsupported validation extensions from XML below;
    # never save the workbook through a library that might discard them.
    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', message='Unknown extension is not supported and will be removed')
        formulas = load_workbook(path, data_only=False, keep_links=False)
        cached = load_workbook(path, data_only=True, keep_links=False)
    raw_sheets = spreadsheet_xml(path)
    try:
        for sheet in formulas:
            meta = {'kind': 'worksheet_metadata', 'sheet': sheet.title, 'visibility': sheet.sheet_state,
                    'document_role': role(path), 'merged_ranges': sorted(str(r) for r in sheet.merged_cells.ranges),
                    'source_xml_part': raw_sheets[sheet.title]['part'],
                    'source_validation_xml': raw_sheets[sheet.title]['validation_xml']}
            yield block(source, f'{sheet.title}!#metadata', f'Worksheet: {sheet.title}; visibility: {sheet.sheet_state}', **meta)
            for row in sheet:
                for cell in row:
                    if cell.value is None and cell.comment is None and cell.hyperlink is None:
                        continue
                    value = cell.value.text if isinstance(cell.value, ArrayFormula) else scalar(cell.value)
                    metadata = {'kind': 'spreadsheet_cell', 'sheet': sheet.title, 'cell': cell.coordinate,
                                'visibility': sheet.sheet_state, 'document_role': role(path),
                                'value': value, 'number_format': cell.number_format, 'data_type': cell.data_type,
                                'source_cell_xml': raw_sheets[sheet.title]['cells'][cell.coordinate]}
                    if isinstance(cell.value, ArrayFormula):
                        metadata['array_range'] = cell.value.ref
                    if cell.data_type == 'f':
                        metadata['formula'] = value
                        metadata['cached_value'] = scalar(cached[sheet.title][cell.coordinate].value)
                        metadata['cache_status'] = 'unavailable' if metadata['cached_value'] is None else 'source_cached_unrecalculated'
                        text = f"Formula: {value}; source cached value: {metadata['cached_value']!r} (not independently recalculated)"
                    else:
                        text = '' if value is None else str(value)
                    if cell.hyperlink:
                        metadata['hyperlink'] = {'target': cell.hyperlink.target, 'location': cell.hyperlink.location}
                    if cell.comment:
                        metadata['comment'] = {'text': cell.comment.text, 'author': cell.comment.author}
                    yield block(source, f'{sheet.title}!{cell.coordinate}', text, **metadata)
    finally:
        formulas.close()
        cached.close()


def spreadsheet_xml(path):
    ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    with ZipFile(path) as archive:
        rels = {r.get('Id'): r.get('Target') for r in ET.fromstring(archive.read('xl/_rels/workbook.xml.rels'))}
        sheets = ET.fromstring(archive.read('xl/workbook.xml')).findall('s:sheets/s:sheet', ns)
        result = {}
        for sheet in sheets:
            target = rels[sheet.get(f"{{{NS['r']}}}id")]
            part = target.lstrip('/') if target.startswith('/') else posixpath.normpath('xl/' + target)
            tree = ET.fromstring(archive.read(part))
            result[sheet.get('name')] = {
                'part': part,
                'cells': {c.get('r'): ET.tostring(c, encoding='unicode') for c in tree.findall('.//s:sheetData/s:row/s:c', ns)},
                'validation_xml': [ET.tostring(e, encoding='unicode') for e in tree
                                   if e.tag.rsplit('}', 1)[-1] in ('dataValidations', 'extLst')],
            }
        return result


def pdf_blocks(path):
    from pypdf import PdfReader
    source = relative(path)
    reader = PdfReader(path)
    for page, content in enumerate(reader.pages, 1):
        text = content.extract_text() or ''
        yield block(source, f'page:{page}', text, kind='pdf_page_text', document_role=role(path),
                    extraction_status='embedded_text' if text.strip() else 'image_only_see_reviewed_visual_transcription')


def read_source(path):
    if path.suffix == '.docx':
        return list(word_blocks(path))
    if path.suffix == '.xlsx':
        return list(sheet_blocks(path))
    if path.suffix == '.pdf':
        return list(pdf_blocks(path))
    return [block(relative(path), 'image', '', kind='source_image', document_role=role(path),
                  extraction_status='see_reviewed_visual_transcription')]


def markdown(blocks):
    """Preserve source text order, line breaks, and table/cell relationships."""
    source = blocks[0]['source']
    out = [f'# {Path(source).stem}', '', f'Source: `{source}`',
           f'Source SHA-256: `{sha(ROOT / source)}`',
           f'Document role: `{blocks[0]["document_role"]}`', '',
           'The text below records source statements. Policy requirements, template text, and requested actions are not proof of implementation or completion.', '']
    for b in blocks:
        context = ''
        if 'table' in b:
            context = f' | table {b["table"]}, XML row {b["row"]}, XML cell {b["cell"]}'
        out.extend([f'<!-- evidence:{b["id"]} source:{b["locator"]}{context} -->', b['text'], ''])
        if b.get('field_instructions'):
            out.extend(['Source field instructions: ' + json.dumps(b['field_instructions'], ensure_ascii=False), ''])
        if b.get('hyperlink_targets') or b.get('hyperlink'):
            out.extend(['Source hyperlinks: ' + json.dumps(b.get('hyperlink_targets', b.get('hyperlink')), ensure_ascii=False), ''])
        if b.get('comment'):
            out.extend(['Source cell comment: ' + json.dumps(b['comment'], ensure_ascii=False), ''])
    return '\n'.join(out)


def extract():
    evidence = ROOT / 'evidence'
    evidence.mkdir(exist_ok=True)
    manifest_path = evidence / 'source_manifest.json'
    current = [{'path': relative(p), 'bytes': p.stat().st_size, 'sha256': sha(p), 'format': p.suffix[1:]} for p in raw_paths()]
    if manifest_path.exists():
        if json.loads(manifest_path.read_text())['sources'] != current:
            raise ValueError('Source manifest mismatch; originals changed. Investigate before updating the baseline.')
    else:
        manifest_path.write_text(json.dumps({'scope': 'Local supplied files at audit start; no comparison with remote Drive was performed.', 'sources': current}, indent=2) + '\n')
    index = json.loads((ROOT / 'corpus_text/00_INDEX/master_corpus_index.json').read_text())
    mappings = {d['source']: d['relative_path'] for d in index['documents'] if d['source'].endswith(('.docx', '.xlsx'))}
    all_blocks = []
    transcriptions = {}
    for path in raw_paths():
        blocks = read_source(path)
        all_blocks.extend(blocks)
        if relative(path) in mappings:
            transcriptions[ROOT / 'corpus_text' / mappings[relative(path)]] = markdown(blocks)
    serialized = ''.join(json.dumps(b, ensure_ascii=False, sort_keys=True, allow_nan=False) + '\n' for b in all_blocks)
    for path, text in transcriptions.items():
        path.write_text(text)
    (evidence / 'blocks.jsonl').write_text(serialized)
    print(f'Extracted {len(all_blocks)} source blocks from {len(current)} unchanged source files')


if __name__ == '__main__':
    extract()
